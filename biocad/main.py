import openpyxl
from openpyxl.chart import BarChart, Reference
import os

from openpyxl.workbook import Workbook


def merge_excel_tables(file_paths):
    merged_data = []
    flag = False
    wb = Workbook()
    ws = wb.active
    timing_data = {}

    for file in os.listdir(file_paths):
        if file.endswith('xlsx'):
            workbook = openpyxl.load_workbook(file_paths + '\\' + file)
            sheet = workbook.active
            table_name = sheet['A1'].value
            merged_data.append([table_name])

            for row in sheet.iter_rows(values_only=True):
                if row[0] is None:
                    break
                elif row[0] == table_name:
                    continue
                elif row[0] == "Задача" and flag:
                    continue
                else:
                    merged_data.append(list(row))
                    flag = True
                    timing_data.update({row[0]: [row[1], row[2]]})

    return merged_data, timing_data


def add_data_to_excel(merged_data, output_file, timing_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active(title='Diagram', index=0)

    for row in merged_data:
        sheet.append(row)

    chart = BarChart()
    chart.title = 'Заголовок'

    data = Reference(sheet, min_col=2, min_row=3, max_col=3, max_row=13)

    chart.add_data(data, titles_from_data=True)
    sheet.add_chart(chart, 'G4')

    workbook.save(output_file)


if __name__ == '__main__':
    merged_data, timing_data = merge_excel_tables('resource')
    add_data_to_excel(merged_data, 'result.xlsx', timing_data)
